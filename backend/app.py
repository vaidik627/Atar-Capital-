from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import time
import traceback
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .ocr_service import extract_text_from_file
from .extraction import extract_financial_data, load_extracted_data

app = Flask(__name__, static_folder='../')
CORS(app)

# Mock Database
DEALS = {}
DOCUMENTS = {}

# Ensure directories exist
EXTRACTED_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'extracted_data')
os.makedirs(EXTRACTED_DATA_DIR, exist_ok=True)

REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)

# Serve Static Files
@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory(app.static_folder, path)

# --- Persistence Recovery ---
def load_existing_deals():
    """Scans extracted_data directory and repopulates DEALS/DOCUMENTS"""
    if not os.path.exists(EXTRACTED_DATA_DIR):
        return

    print(f"Scanning for existing deals in: {EXTRACTED_DATA_DIR}")
    for filename in os.listdir(EXTRACTED_DATA_DIR):
        if filename.endswith('.json'):
            try:
                deal_id = filename.replace('.json', '')
                filepath = os.path.join(EXTRACTED_DATA_DIR, filename)
                
                with open(filepath, 'r', encoding='utf-8') as f:
                    extracted_data = json.load(f)
                
                # Reconstruct deal info
                company_name = extracted_data.get('company_name') or f"Deal {deal_id}"
                deal_value = extracted_data.get('market_intelligence', {}).get('market_size') or "N/A"
                
                # Populate DEALS
                DEALS[deal_id] = {
                    "id": deal_id,
                    "name": company_name,
                    "value": deal_value,
                    "status": "Active",
                    "date": time.strftime("%Y-%m-%d", time.localtime(os.path.getmtime(filepath))),
                    "file_name": f"{deal_id}.pdf" # Approximate
                }
                
                # Populate DOCUMENTS (Mock)
                doc_id = f"doc_{deal_id}"
                DOCUMENTS[doc_id] = {
                    "id": doc_id,
                    "deal_id": deal_id,
                    "name": f"{deal_id}.pdf",
                    "extracted_data": extracted_data,
                    "ocr_text_preview": "Loaded from disk..."
                }
                print(f"Recovered deal: {deal_id} ({company_name})")
                
            except Exception as e:
                print(f"Failed to load deal from {filename}: {e}")

# Load deals on startup
load_existing_deals()

# --- API Endpoints ---

@app.route('/api/deals', methods=['GET', 'POST'])
def handle_deals():
    if request.method == 'POST':
        data = request.json
        deal_id = str(len(DEALS) + 1)
        DEALS[deal_id] = {
            "id": deal_id,
            "name": data.get('name', 'New Deal'),
            "status": "Active",
            "date": "2024-10-25"
        }
        return jsonify({"success": True, "deal": DEALS[deal_id]})
    return jsonify({"success": True, "deals": list(DEALS.values())})

@app.route('/api/documents', methods=['GET'])
def list_documents():
    deal_id = request.args.get('dealId')
    docs = [doc for doc in DOCUMENTS.values() if doc['deal_id'] == deal_id] if deal_id else list(DOCUMENTS.values())
    return jsonify({
        "success": True,
        "data": docs
    })

@app.route('/api/analysis/<deal_id>', methods=['GET'])
def get_analysis(deal_id):
    extracted_data = None
    
    # 1. Try to find in memory first
    deal_docs = [doc for doc in DOCUMENTS.values() if doc['deal_id'] == deal_id]
    if deal_docs:
        extracted_data = deal_docs[-1]['extracted_data']
    
    # 2. If not in memory, try to load from disk
    if not extracted_data:
        extracted_data = load_extracted_data(deal_id)
        
    if not extracted_data:
        return jsonify({
            "success": False,
            "message": "Analysis not found",
            "data": {}
        })
    
    # Format for frontend
    # Map the backend schema to the frontend expected structure
    # Frontend expects: header, revenue, profitMetrics, marketIntelligence, riskAnalysis, aiSuggestion
    
    # Retrieve user deal value from DEALS dictionary
    user_deal_value = "N/A"
    if deal_id in DEALS:
        user_deal_value = DEALS[deal_id].get("value", "N/A")

    analysis_data = {
        "header": {
            "companyName": extracted_data.get('company_name'),
            "currency": extracted_data.get('currency', 'USD'),
            "dealValue": {
                "description": "Deal Value",
                "display": user_deal_value,
                "visible": True
            },
            "marketSize": {
                "description": "Market Size",
                "display": extracted_data.get('market_intelligence', {}).get('market_size') or "N/A",
                "visible": True
            }
        },
        "summary": {
            "text": extracted_data.get('company_summary', {}).get('text', ''),
            "visible": True
        },
        "revenue": {
            "present": {
                "value": extracted_data.get('revenue', {}).get('present', {}).get('value', 'N/A'),
                "period": extracted_data.get('revenue', {}).get('present', {}).get('period', 'N/A')
            },
            "history": extracted_data.get('revenue', {}).get('history', []),
            "future": extracted_data.get('revenue', {}).get('future', []),
            "visible": True
        },
        "profitMetrics": extracted_data.get('profit_metrics', {}),
        "marketIntelligence": {
            "industryPosition": extracted_data.get('market_intelligence', {}).get('industry_position'),
            "marketSharePercent": extracted_data.get('market_intelligence', {}).get('market_share_percent'),
            "competitors": extracted_data.get('market_intelligence', {}).get('key_competitors', []),
            "context": extracted_data.get('market_intelligence', {}).get('source_context'),
            "visible": True
        },
        "riskAnalysis": {
            "operational": extracted_data.get('risk_analysis', {}).get('operational_risks', []),
            "financial": extracted_data.get('risk_analysis', {}).get('financial_risks', []),
            "market": extracted_data.get('risk_analysis', {}).get('market_risks', []),
            "regulatory": extracted_data.get('risk_analysis', {}).get('regulatory_risks', []),
            "visible": True
        },
        "aiSuggestion": {
            "recommendation": extracted_data.get('ai_suggestion', {}).get('recommendation'),
            "confidence": extracted_data.get('ai_suggestion', {}).get('confidence_percent'),
            "rationale": extracted_data.get('ai_suggestion', {}).get('rationale'),
            "visible": True
        }
    }

    return jsonify({
        "success": True,
        "data": analysis_data
    })

@app.route('/api/extract', methods=['POST'])
@app.route('/api/documents/upload', methods=['POST'])
def process_document():
    deal_id = request.form.get('dealId') # Frontend sends dealId in FormData
    deal_name = request.form.get('dealName')
    deal_value = request.form.get('dealValue')
    
    if 'file' not in request.files:
        # Check for 'document' or 'documents' field as well for compatibility
        if 'document' in request.files:
            file = request.files['document']
        elif 'documents' in request.files:
            file = request.files['documents']
        else:
            return jsonify({"success": False, "message": "No file part"}), 400
    else:
        file = request.files['file']
    
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400
    
    try:
        # Generate deal_id if not provided
        if not deal_id:
            deal_id = f"deal_{int(time.time())}_{int(os.urandom(4).hex(), 16)}"
            
        # Register/Update Deal in Mock DB
        DEALS[deal_id] = {
            "id": deal_id,
            "name": deal_name if deal_name else f"Deal {deal_id}",
            "value": deal_value if deal_value else "N/A",
            "status": "Processing",
            "date": time.strftime("%Y-%m-%d"),
            "file_name": file.filename
        }
        
        # 1. OCR Extraction
        file_content = file.read()
        mime_type = file.mimetype or 'application/pdf'
        print(f"Processing file: {file.filename} ({mime_type}) for Deal: {deal_id}")
        
        ocr_text = extract_text_from_file(file_content, mime_type)
        print("OCR complete. Text length:", len(ocr_text))

        # Save OCR text to backend/parsed_text
        parsed_text_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'parsed_text')
        os.makedirs(parsed_text_dir, exist_ok=True)
        
        ocr_filename = f"{int(time.time())}_{file.filename}.txt"
        ocr_filepath = os.path.join(parsed_text_dir, ocr_filename)
        
        with open(ocr_filepath, 'w', encoding='utf-8') as f:
            f.write(ocr_text)
        print(f"OCR text saved to: {ocr_filepath}")
        
        # 2. Financial Data Extraction
        # Use deal_id from request if available, otherwise construct one
        extraction_deal_id = deal_id if deal_id else f"upload_{file.filename}"
        extracted_data = extract_financial_data(ocr_text, deal_id=extraction_deal_id, user_deal_value=deal_value)
        
        # Store result (Mock)
        doc_id = f"doc_{int(time.time())}_{file.filename}"
        DOCUMENTS[doc_id] = {
            "id": doc_id,
            "deal_id": extraction_deal_id,
            "name": file.filename,
            "extracted_data": extracted_data,
            "ocr_text_preview": ocr_text[:500]
        }
        
        # Update Deal Status
        if extraction_deal_id in DEALS:
            DEALS[extraction_deal_id]["status"] = "Active"

        return jsonify({
            "success": True,
            "dealId": extraction_deal_id,
            "data": extracted_data,
            "ocr_text_preview": ocr_text[:500]
        })
        
    except Exception as e:
        print(f"Error processing document: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "healthy", "service": "Financial Extraction Engine"})

@app.route('/api/reports/generate', methods=['POST'])
def generate_report():
    try:
        data = request.json
        deal_id = data.get('dealId')
        
        # Initialize Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pre-Bid Analysis"
        
        # --- Styles ---
        bold_font = Font(bold=True)
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center')
        left_align = Alignment(horizontal='left')
        right_align = Alignment(horizontal='right')
        
        # Red fill for Atar Projections (light red/pink)
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        # Blue fill for Management Projections (light blue)
        blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        # Grey fill for headers
        grey_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # --- Section 1: Inputs (Left Panel) ---
        
        # Title
        ws['D2'] = "Project Manta Ray Pre-Bid Analysis"
        ws['D2'].font = title_font
        ws['D3'] = deal_id
        
        # Base Inputs
        ws['A4'] = "Base Inputs"
        ws['A4'].font = bold_font
        
        base_inputs = [
            "WC Availability",
            "Debt Sourcing Fees",
            "Diligence Fees",
            "Atar Transaction Fees",
            "FCCR",
            "SOFR"
        ]
        
        for i, label in enumerate(base_inputs, start=5):
            ws[f'A{i}'] = label
            # ws[f'B{i}'] = val  <-- Removed value
            ws[f'B{i}'].alignment = center_align
            ws[f'B{i}'].fill = blue_fill
            
        # Financing Inputs
        ws['A12'] = "Financing Inputs"
        ws['A12'].font = bold_font
        
        financing_inputs = [
            "Equity",
            "Revolver",
            "Revolver",
            "Term Loan",
            "Term Loan",
            "Structure",
            "Structure",
            "Structure"
        ]
        
        for i, label in enumerate(financing_inputs, start=13):
            ws[f'A{i}'] = label
            # ws[f'B{i}'] = val <-- Removed value
            # ws[f'C{i}'] = pct <-- Removed value
            
        # Financial Projections Selection
        ws['A24'] = "Financial Projections"
        ws['A24'].font = bold_font
        ws['B24'] = "Atar"
        ws['B24'].fill = blue_fill
        
        # Purchase Assumptions
        ws['B26'] = "Purchase Assumptions"
        ws['B26'].font = bold_font
        ws['B27'] = "EBITDA"
        ws['C27'] = "Multiple"
        # ws['B28'] = "$23,000" <-- Removed value
        # ws['C28'] = "3.0x"    <-- Removed value
        
        # Exit Assumptions
        ws['B32'] = "Exit Assumptions"
        ws['B32'].font = bold_font
        ws['B33'] = "EBITDA"
        ws['C33'] = "Multiple"
        # ws['B34'] = "$35,000" <-- Removed value
        # ws['C34'] = "3.0x"    <-- Removed value
        
        # Shareholder Returns
        ws['D32'] = "Shareholder Returns"
        ws['D32'].font = bold_font
        
        shareholder_rows = [
            "Exit",
            "EV @ Exit",
            "+ Cash",
            "- Debt",
            "- Expenses",
            "- Mgmt LTIP",
            "- Seller Equity",
            "Atar Equity"
        ]
        
        for i, label in enumerate(shareholder_rows, start=33):
            ws[f'D{i}'] = label
            # ws[f'F{i}'] = val <-- Removed value
        
        # Distribution of Proceeds
        ws['D42'] = "Distribution of Proceeds"
        ws['D42'].font = bold_font
        
        dist_rows = [
            "Return of Equity",
            "Interest on Preferred Shares",
            "LP Split of Proceeds",
            "LP Total Distribution",
            "LP MOIC",
            "GP Split of Proceeds"
        ]
        
        for i, label in enumerate(dist_rows, start=43):
            ws[f'D{i}'] = label
            # ws[f'F{i}'] = val <-- Removed value
            
        # --- Section 2: Financial Model (Right Panel) ---
        
        # Columns Configuration
        # K: Labels
        # L-O: Actuals (2019A-2022A)
        # P-T: Atar Projections (2023 R - 2027 R)
        # U-Y: Mgmt Projections (2023 M - 2027 M)
        
        # Headers
        ws['K2'] = "Financials"
        ws['K2'].font = bold_font
        
        ws.merge_cells('L2:O2')
        ws['L2'] = "Actual"
        ws['L2'].alignment = center_align
        ws['L2'].font = bold_font
        
        ws.merge_cells('P2:T2')
        ws['P2'] = "Atar Projections"
        ws['P2'].alignment = center_align
        ws['P2'].font = bold_font
        ws['P2'].fill = red_fill
        
        ws.merge_cells('U2:Y2')
        ws['U2'] = "Management Projections"
        ws['U2'].alignment = center_align
        ws['U2'].font = bold_font
        ws['U2'].fill = blue_fill
        
        # Years Row
        years_actual = ["2019A", "2020A", "2021A", "2022A"]
        years_atar = ["2023 R", "2024 R", "2025 R", "2026 R", "2027 R"]
        years_mgmt = ["2023 M", "2024 M", "2025 M", "2026 M", "2027 M"]
        
        col_idx = 12 # Column L
        for y in years_actual:
            cell = ws.cell(row=3, column=col_idx, value=y)
            cell.alignment = center_align
            cell.font = bold_font
            col_idx += 1
            
        for y in years_atar:
            cell = ws.cell(row=3, column=col_idx, value=y)
            cell.alignment = center_align
            cell.font = bold_font
            cell.fill = red_fill
            col_idx += 1
            
        for y in years_mgmt:
            cell = ws.cell(row=3, column=col_idx, value=y)
            cell.alignment = center_align
            cell.font = bold_font
            cell.fill = blue_fill
            col_idx += 1
            
        # Financial Rows
        financial_labels = [
            "Net Revenue",
            "% Growth",
            "",
            "Gross Profit",
            "% Margin",
            "",
            "Operating Expenses",
            "% Growth",
            "",
            "Reported EBITDA",
            "Adjustments",
            "Adj. EBITDA",
            "% Margin"
        ]
        
        start_row = 4
        for i, label in enumerate(financial_labels):
            ws.cell(row=start_row + i, column=11, value=label) # Column K
            
        # Tale of The Tape
        tape_start = start_row + len(financial_labels) + 2
        ws.cell(row=tape_start, column=11, value="Tale of The Tape").font = bold_font
        
        tape_labels = [
            "Adj. EBITDA",
            "Capex",
            "Change in WC",
            "1x Costs",
            "Free Cash Flow"
        ]
        
        for i, label in enumerate(tape_labels):
            ws.cell(row=tape_start + 1 + i, column=11, value=label)
            
        # Interest
        interest_start = tape_start + len(tape_labels) + 2
        ws.cell(row=interest_start, column=11, value="Interest").font = bold_font
        
        interest_labels = [
            "Revolver",
            "Term Loan",
            "Seller Note",
            "Interest Subtotal"
        ]
        
        for i, label in enumerate(interest_labels):
            ws.cell(row=interest_start + 1 + i, column=11, value=label)
            
        # Amortization
        amort_start = interest_start + len(interest_labels) + 2
        ws.cell(row=amort_start, column=11, value="Amortization").font = bold_font
        
        amort_labels = [
            "Term Loan",
            "Seller Note",
            "Amortization Subtotal"
        ]
        
        for i, label in enumerate(amort_labels):
            ws.cell(row=amort_start + 1 + i, column=11, value=label)
            
        # Summary Metrics
        summary_start = amort_start + len(amort_labels) + 2
        ws.cell(row=summary_start, column=11, value="Total Debt Service").font = bold_font
        ws.cell(row=summary_start+2, column=11, value="FCF After Debt Service").font = bold_font
        ws.cell(row=summary_start+3, column=11, value="Management Fees")
        ws.cell(row=summary_start+4, column=11, value="Earnout")
        
        revolver_start = summary_start + 6
        ws.cell(row=revolver_start, column=11, value="Cash Available for Revolver").font = bold_font
        ws.cell(row=revolver_start+1, column=11, value="Revolver Drawdown (Repayment)")
        ws.cell(row=revolver_start+2, column=11, value="Revolver Balance")
        ws.cell(row=revolver_start+3, column=11, value="Remaining Cash")
        
        fccr_start = revolver_start + 5
        ws.cell(row=fccr_start, column=11, value="FCCR - Mgmt fee is a fixed charge")
        ws.cell(row=fccr_start+1, column=11, value="FCCR - Mgmt fee is NOT a fixed charge")
        
        # Column Width Adjustments
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['K'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']:
             ws.column_dimensions[col].width = 12

        filename = f"report_{deal_id}_{int(time.time())}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        wb.save(filepath)
        
        return jsonify({"success": True, "filename": filename})
    except Exception as e:
        print(f"Error generating report: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/reports', methods=['GET'])
def list_reports():
    reports = []
    if os.path.exists(REPORTS_DIR):
        for f in os.listdir(REPORTS_DIR):
            if f.endswith('.xlsx'):
                reports.append({"filename": f, "created_at": os.path.getctime(os.path.join(REPORTS_DIR, f))})
    return jsonify({"success": True, "reports": sorted(reports, key=lambda x: x['created_at'], reverse=True)})

@app.route('/api/reports/download/<filename>', methods=['GET'])
def download_report(filename):
    return send_from_directory(REPORTS_DIR, filename, as_attachment=True)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8000))
    app.run(host='0.0.0.0', port=port, debug=True)
