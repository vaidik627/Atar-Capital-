import os
import csv
import time
import json
import re
from typing import Any, Dict, Optional, Tuple, List
from .config import REPORTS_DIR, get_excel_template_path

def generate_csv_report(deal_id, deal_name, data):
    """
    Generates a CSV report for the deal.
    Exports Revenue, Profit Metrics, Market Intelligence, and Risks.
    
    Returns:
        str: The filename of the generated report.
    """
    
    timestamp = int(time.time())
    # Clean deal name for filename
    safe_deal_name = "".join([c for c in deal_name if c.isalnum() or c in (' ', '-', '_')]).strip().replace(' ', '_')
    filename = f"{safe_deal_name}_{deal_id}_Analysis_{timestamp}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    # Define headers
    header = ["Category", "Metric", "Period", "Value", "Unit/Currency"]
    
    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header)
        
        # 1. Revenue
        revenue = data.get('revenue', {})
        
        # Present
        present = revenue.get('present', {})
        if present:
            writer.writerow(["Revenue", "Present Revenue", present.get('period', 'N/A'), present.get('value', 'N/A'), ""])
            
        # History
        for item in revenue.get('history', []):
            writer.writerow(["Revenue", "Historical Revenue", item.get('period', 'N/A'), item.get('value', 'N/A'), item.get('unit', '')])
            
        # Future
        for item in revenue.get('future', []):
            writer.writerow(["Revenue", "Projected Revenue", item.get('period', 'N/A'), item.get('value', 'N/A'), item.get('unit', '')])
            
        # 2. Profit Metrics
        profit = data.get('profit_metrics', {})
        for metric, items in profit.items():
            metric_name = metric.replace('_', ' ').title()
            for item in items:
                writer.writerow(["Profit Metrics", metric_name, item.get('period', 'N/A'), item.get('value', 'N/A'), item.get('unit', '')])
                
        # 3. Market Intelligence
        market = data.get('market_intelligence', {})
        if market.get('market_size'):
             writer.writerow(["Market", "Market Size", "", market.get('market_size'), ""])
        if market.get('industry_position'):
             writer.writerow(["Market", "Industry Position", "", market.get('industry_position'), ""])
             
        # 4. Risks
        risks = data.get('risk_analysis', {})
        for r_type, items in risks.items():
            if isinstance(items, list):
                type_name = r_type.replace('_', ' ').title()
                for item in items:
                    writer.writerow(["Risk", type_name, "", item, ""])
        
    return filename


def generate_excel_report(deal_id: str, deal_name: str, data: Dict[str, Any], template_path: Optional[str] = None) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(f"openpyxl is required for Excel export: {e}")

    timestamp = int(time.time())
    safe_deal_name = "".join([c for c in deal_name if c.isalnum() or c in (' ', '-', '_')]).strip().replace(' ', '_')
    filename = f"{safe_deal_name}_{deal_id}_Model_{timestamp}.xlsm"
    filepath = os.path.join(REPORTS_DIR, filename)

    template_path = template_path or get_excel_template_path()
    # Load with data_only=True so openpyxl reads cached cell values, not formula strings.
    # This prevents #DIV/0! / formula strings from leaking into our output.
    wb = load_workbook(template_path, keep_vba=True, data_only=True)

    updated_any = False
    base_year = _infer_base_year(data)
    for ws in _pick_target_sheets(wb):
        # print(f"DEBUG: Checking sheet '{ws.title}'")
        unit_scale = _detect_template_unit_scale(ws)
        year_blocks = _detect_year_blocks(ws)
        if not year_blocks:
            # print(f"DEBUG: No year blocks found in '{ws.title}'")
            continue
        # print(f"DEBUG: Found {len(year_blocks)} year blocks in '{ws.title}'")

        row_map = _detect_row_map(ws)
        if not any(isinstance(v, int) for v in row_map.values()):
            if _sheet_contains_text(ws, "Assets") and _sheet_contains_text(ws, "Liabilities"):
                bs_data = data.get("balance_sheet")
                if bs_data:
                    _write_balance_sheet_to_excel(ws, year_blocks, bs_data, unit_scale)
                    updated_any = True
            continue
        # print(f"DEBUG: Row map found: {row_map}")

        _update_template_header(ws, deal_name, data.get("currency"))

        from typing import List

        revenue_by_year = _extract_series_by_year(_collect_revenue_items(data), unit_scale)
        actual_years_sorted = _get_actual_years_from_data(data, base_year)
        if actual_years_sorted:
            base_year = actual_years_sorted[-1]


        for block in year_blocks:
            _rewrite_year_header_row(ws, block, base_year, actual_years_sorted)
        year_blocks = _detect_year_blocks(ws)
        _clear_template_inputs(ws, year_blocks, row_map)

        gross_profit_by_year = _extract_series_by_year(data.get("profit_metrics", {}).get("gross_profit", []), unit_scale)
        ebitda_by_year = _extract_series_by_year(data.get("profit_metrics", {}).get("ebitda", []), unit_scale)
        adj_ebitda_by_year = _extract_series_by_year(data.get("profit_metrics", {}).get("adjusted_ebitda", []), unit_scale)

        # Create separate projection sets for ATAR (AI) and Management
        # ATAR: Conservative/Base AI projection
        # Management: Optimistic/Management AI projection (if missing)
        
        revenue_atar = revenue_by_year.copy()
        gross_profit_atar = gross_profit_by_year.copy()
        ebitda_atar = ebitda_by_year.copy()
        
        revenue_mgmt = revenue_by_year.copy()
        gross_profit_mgmt = gross_profit_by_year.copy()
        ebitda_mgmt = ebitda_by_year.copy()

        operating_income_by_year = _extract_series_by_year(
            data.get("profit_metrics", {}).get("operating_income", []),
            unit_scale,
        )
        
        operating_expense_by_year = _extract_series_by_year(
            data.get("profit_metrics", {}).get("operating_expenses", []),
            unit_scale,
        )
        
        # OpEx for Actual -> MUST be strictly extracted Operating Expenses (no derivation allowed)
        opex_actual = operating_expense_by_year

        # OpEx for ATAR -> Fallback to derived if missing for projections
        opex_atar = _derive_opex(gross_profit_atar, operating_income_by_year)
        if gross_profit_atar and ebitda_atar:
            for y in set(gross_profit_atar.keys()) & set(ebitda_atar.keys()):
                if y not in opex_atar:
                    opex_atar[y] = gross_profit_atar[y] - ebitda_atar[y]
                    
        # OpEx for Management
        opex_mgmt = _derive_opex(gross_profit_mgmt, operating_income_by_year)
        if gross_profit_mgmt and ebitda_mgmt:
            for y in set(gross_profit_mgmt.keys()) & set(ebitda_mgmt.keys()):
                if y not in opex_mgmt:
                    opex_mgmt[y] = gross_profit_mgmt[y] - ebitda_mgmt[y]

        # Fill projections with different assumptions
        _fill_future_projections(
            base_year, 
            revenue_atar, 
            gross_profit_atar, 
            operating_income_by_year.copy(),
            opex_atar,
            ebitda_atar, 
            mode="atar"
        )
        _fill_future_projections(
            base_year, 
            revenue_mgmt, 
            gross_profit_mgmt, 
            operating_income_by_year.copy(),
            opex_mgmt,
            ebitda_mgmt, 
            mode="management"
        )

        # Derive secondary metrics for both scenarios
        cogs_atar = _derive_cogs(revenue_atar, gross_profit_atar)
        cogs_mgmt = _derive_cogs(revenue_mgmt, gross_profit_mgmt)
        
        capex_by_year = _extract_tale_year_wise(data, "capex", unit_scale)
        wc_by_year = _extract_tale_year_wise(data, "change_in_working_capital", unit_scale)
        one_time_by_year = _extract_tale_year_wise(data, "one_time_cost", unit_scale)

        fcf_hist_by_year, fcf_forecast_by_year = _extract_fcf(data, unit_scale)

        if not adj_ebitda_by_year:
            adj_ebitda_by_year = _derive_adj_ebitda(ebitda_by_year, one_time_by_year)
            
        adj_ebitda_atar = _derive_adj_ebitda(ebitda_atar, one_time_by_year)
        adj_ebitda_mgmt = _derive_adj_ebitda(ebitda_mgmt, one_time_by_year)

        fcf_all = dict(fcf_hist_by_year)
        capex_atar, capex_mgmt = _prepare_tale_projections(base_year, capex_by_year, revenue_atar, revenue_mgmt, "capex")
        wc_atar, wc_mgmt = _prepare_tale_projections(base_year, wc_by_year, revenue_atar, revenue_mgmt, "wc")
        one_time_atar, one_time_mgmt = _prepare_tale_projections(base_year, one_time_by_year, revenue_atar, revenue_mgmt, "one_time")

        # 1. Get dynamic projection years for Atar explicitly
        projection_years = list(range(base_year + 1, base_year + 6))
        
        atar_blocks = _detect_all_atar_blocks(ws)
        if not atar_blocks:
            for b in year_blocks:
                if b.get("role") == "projection" and b.get("year_cols"):
                    yc = b["year_cols"]
                    atar_blocks.append({
                        "header_row": b.get("header_row", 5),
                        "start_col": min(yc.values()),
                        "end_col": max(yc.values()),
                        "role": "projection"
                    })

        for block in year_blocks:
            role = block.get("role", "projection")
            if role == "projection": 
                continue # Skip generic projection handling here, handled securely by atar_blocks iteration below
                
            year_cols = block.get("year_cols", {})
            if not isinstance(year_cols, dict) or not year_cols:
                continue
            
            # Select appropriate dataset
            if role == "management":
                rev_set = revenue_mgmt
                gp_set = gross_profit_mgmt
                ebitda_set = ebitda_mgmt
                adj_ebitda_set = adj_ebitda_mgmt
                cogs_set = cogs_mgmt
                opex_set = opex_mgmt
                capex_set = capex_mgmt
                wc_set = wc_mgmt
                one_time_set = one_time_mgmt
            elif role == "actual":
                rev_set = revenue_by_year
                gp_set = gross_profit_by_year
                ebitda_set = ebitda_by_year
                adj_ebitda_set = adj_ebitda_by_year
                cogs_set = _derive_cogs(revenue_by_year, gross_profit_by_year)
                opex_set = opex_actual
                capex_set = capex_by_year
                wc_set = wc_by_year
                one_time_set = one_time_by_year

            _write_line(ws, row_map.get("net_revenue"), year_cols, rev_set, template_scale=unit_scale)
            _write_line(ws, row_map.get("cogs"), year_cols, cogs_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("gross_profit"), year_cols, gp_set, template_scale=unit_scale)
            _write_line(ws, row_map.get("operating_expense"), year_cols, opex_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("ebitda"), year_cols, ebitda_set, is_ebitda=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("pf_adjustments"), year_cols, one_time_set, template_scale=unit_scale)
            _write_line(ws, row_map.get("adj_ebitda"), year_cols, adj_ebitda_set, is_ebitda=True, template_scale=unit_scale)
            
            # Tale of the Tape - populate based on role
            _write_line(ws, row_map.get("capex"), year_cols, capex_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("change_in_wc"), year_cols, wc_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("one_time_cost"), year_cols, one_time_set, force_negative=False, template_scale=unit_scale)
            # FCF row is typically a formula (=SUM), so we skip it or write only if no formula exists
            _write_line(ws, row_map.get("free_cash_flow"), year_cols, fcf_all if role == "actual" else {}, force_negative=False, template_scale=unit_scale)
            updated_any = True
            
        # 2. Process all explicitly detected Atar Projection blocks exactly as requested over the dynamically mapped coordinates.
        for block in atar_blocks:
            n_cols = block["end_col"] - block["start_col"] + 1
            year_cols = {}
            for i in range(n_cols):
                y = projection_years[i] if i < len(projection_years) else projection_years[-1] + (i - len(projection_years) + 1)
                year_cols[y] = block["start_col"] + i
                
            rev_set = revenue_atar
            gp_set = gross_profit_atar
            ebitda_set = ebitda_atar
            adj_ebitda_set = adj_ebitda_atar
            cogs_set = cogs_atar
            opex_set = opex_atar
            capex_set = capex_atar
            wc_set = wc_atar
            one_time_set = one_time_atar
            
            _write_line(ws, row_map.get("net_revenue"), year_cols, rev_set, template_scale=unit_scale)
            _write_line(ws, row_map.get("cogs"), year_cols, cogs_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("gross_profit"), year_cols, gp_set, template_scale=unit_scale)
            _write_line(ws, row_map.get("operating_expense"), year_cols, opex_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("ebitda"), year_cols, ebitda_set, is_ebitda=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("pf_adjustments"), year_cols, one_time_set, template_scale=unit_scale)
            _write_line(ws, row_map.get("adj_ebitda"), year_cols, adj_ebitda_set, is_ebitda=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("capex"), year_cols, capex_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("change_in_wc"), year_cols, wc_set, force_negative=True, template_scale=unit_scale)
            _write_line(ws, row_map.get("one_time_cost"), year_cols, one_time_set, force_negative=False, template_scale=unit_scale)
            updated_any = True

        # ── 3. FCF + Debt model (pre-computed in Python, no Excel formulas) ──────
        actual_year_cols: Dict[int, int] = {}
        proj_year_cols: Dict[int, int] = {}
        for block in year_blocks:
            role_b = block.get("role", "projection")
            yc = block.get("year_cols", {})
            if role_b == "actual":
                actual_year_cols.update(yc)

        if atar_blocks:
            ab0 = atar_blocks[0]
            n0 = ab0["end_col"] - ab0["start_col"] + 1
            for i in range(n0):
                y = projection_years[i] if i < len(projection_years) else projection_years[-1] + (i - len(projection_years) + 1)
                proj_year_cols[y] = ab0["start_col"] + i

        actual_fcf_model = _compute_fcf_model(
            years=sorted(actual_year_cols.keys()),
            adj_ebitda=adj_ebitda_by_year,
            capex=capex_by_year,
            change_wc=wc_by_year,
            one_time=one_time_by_year,
            interest={},
            amortization={},
            is_actual=True,
            fcf_hist=fcf_hist_by_year
        )
        _write_fcf_model_to_excel(ws, row_map, actual_year_cols, actual_fcf_model, template_scale=unit_scale)

        proj_fcf_model = _compute_fcf_model(
            years=projection_years,
            adj_ebitda=adj_ebitda_atar,
            capex=capex_atar,
            change_wc=wc_atar,
            one_time=one_time_atar,
            interest={},
            amortization={},
            is_actual=False
        )
        for ab in atar_blocks:
            n_ab = ab["end_col"] - ab["start_col"] + 1
            ab_yc: Dict[int, int] = {}
            for i in range(n_ab):
                y = projection_years[i] if i < len(projection_years) else projection_years[-1] + (i - len(projection_years) + 1)
                ab_yc[y] = ab["start_col"] + i
            _write_fcf_model_to_excel(ws, row_map, ab_yc, proj_fcf_model, template_scale=unit_scale)

        # ── 3.5 Inject Default Deal Assumptions ────────
        if _sheet_contains_text(ws, "Purchase Assumptions") or _sheet_contains_text(ws, "Exit Assumptions"):
            _inject_default_assumptions(ws, data)
            _inject_debt_profile(ws, data)

        # ── 4. Final pass: erase any remaining #DIV/0! / formula errors ────────
        _erase_excel_errors(ws)

    # Final global pass: erase errors on ALL sheets (handles cross-sheet formula refs)
    for ws_all in wb.worksheets:
        _erase_excel_errors(ws_all)

    if not updated_any:
        raise ValueError("Could not locate target sheets/rows to populate in the template.")

    wb.save(filepath)
    return filename


def _safe_float(v: Any) -> Optional[float]:
    """Return float or None – never NaN/inf."""
    if v is None:
        return None
    try:
        f = float(v)
        if f != f or f == float("inf") or f == float("-inf"):
            return None
        return f
    except (TypeError, ValueError):
        return None


def _compute_fcf_model(
    years: List[int],
    adj_ebitda: Dict[int, float],
    capex: Dict[int, float],
    change_wc: Dict[int, float],
    one_time: Dict[int, float],
    interest: Dict[int, float] = None,
    amortization: Dict[int, float] = None,
    is_actual: bool = False,
    fcf_hist: Dict[int, float] = None
) -> Dict[int, Dict[str, Optional[float]]]:
    """
    Compute the FCF / Debt model fully in Python for each year.
    Returns a dict: { year: { field: float|None } }
    """
    interest = interest or {}
    amortization = amortization or {}
    fcf_hist = fcf_hist or {}
    model: Dict[int, Dict[str, Optional[float]]] = {}
    revolver_balance_prior = 0.0

    for y in sorted(years):
        ae = _safe_float(adj_ebitda.get(y))
        cx = _safe_float(capex.get(y))
        wc = _safe_float(change_wc.get(y))
        ot = _safe_float(one_time.get(y))
        inte = _safe_float(interest.get(y))
        amor = _safe_float(amortization.get(y))

        if is_actual:
            # STEP 2 & 4: Actuals logic
            # Use extracted FCF for actuals; DO NOT recalculate unless inputs exist.
            if y in fcf_hist:
                fcf = _safe_float(fcf_hist[y])
            else:
                fcf = None
                
            interest_total = inte
            amort_total = amor
            if interest_total is not None or amort_total is not None:
                total_ds = (interest_total or 0.0) + (amort_total or 0.0)
            else:
                total_ds = None

            if fcf is not None and total_ds is not None:
                fcf_after_ds = fcf - total_ds
            elif fcf is not None:
                fcf_after_ds = fcf
            else:
                fcf_after_ds = None

            draw = None
            revolver_balance = None
            remaining = None

        else:
            # STEP 3 & 8: Projection logic
            adj_val = ae if ae is not None else 0.0
            cx_val = cx if cx is not None else 0.0
            wc_val = wc if wc is not None else 0.0
            ot_val = ot if ot is not None else 0.0

            # Step 3: FCF calculation for projections
            fcf = adj_val - abs(cx_val) - wc_val - abs(ot_val)

            # Step 4-5: Debt calculation
            interest_total = inte if inte is not None else 0.0
            amort_total = amor if amor is not None else 0.0
            total_ds = interest_total + amort_total

            # Step 6: FCF after debt
            fcf_after_ds = fcf - total_ds

            # Step 7: Revolver logic
            if fcf_after_ds < 0:
                draw = abs(fcf_after_ds)
                remaining = 0.0
            else:
                draw = 0.0
                remaining = fcf_after_ds

            revolver_balance = revolver_balance_prior + draw
            revolver_balance_prior = revolver_balance

        model[y] = {
            "free_cash_flow":     fcf,
            "interest":           interest_total,
            "amortization":       amort_total,
            "total_debt_service": total_ds,
            "fcf_after_debt":     fcf_after_ds,
            "cash_avail_revolver": fcf_after_ds if not is_actual else None,
            "revolver_draw":      draw,
            "revolver_balance":   revolver_balance,
            "remaining_cash":     remaining,
        }

    return model


def _write_fcf_model_to_excel(
    ws,
    row_map: Dict[str, Optional[int]],
    year_cols: Dict[int, int],
    model: Dict[int, Dict[str, Optional[float]]],
    template_scale: Optional[float] = None
) -> None:
    """Write pre-computed FCF/Debt model values to Excel cells. No formulas."""
    field_keys = [
        "free_cash_flow",
        "interest",
        "amortization",
        "total_debt_service",
        "fcf_after_debt",
        "cash_avail_revolver",
        "revolver_draw",
        "revolver_balance",
        "remaining_cash",
    ]
    for field in field_keys:
        row = row_map.get(field)
        if row is None:
            continue
        for y, col in year_cols.items():
            year_data = model.get(y)
            if year_data is None:
                ws.cell(row=row, column=col).value = "-"
                continue
            v = _safe_float(year_data.get(field))
            cell = ws.cell(row=row, column=col)
            if v is None:
                cell.value = "-"
            else:
                cell.value = v
                if template_scale is not None:
                    if template_scale >= 1_000_000.0:
                        fmt = '"$"#,##0.0"M";[Red]("$"#,##0.0"M")'
                    elif template_scale >= 1_000.0:
                        fmt = '"$"#,##0.0,"M";[Red]("$"#,##0.0,"M")'
                    else:
                        fmt = '"$"#,##0.0,,"M";[Red]("$"#,##0.0,,"M")'
                    cell.number_format = fmt


def _update_template_header(ws, deal_name: str, currency: Any) -> None:
    deal_name = str(deal_name or "").strip()
    if not deal_name:
        return

    target_placeholders = {
        "herff jones",
        "company name",
        "deal name",
        "project name",
        "client name",
        "target name",
        "[company name]",
        "<company name>"
    }

    replaced = False
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), min_col=1, max_col=min(ws.max_column, 20)):
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                s = v.strip().lower()
                if s in target_placeholders:
                    cell.value = deal_name
                    replaced = True
    if replaced:
        return


def _inject_default_assumptions(ws, data: Dict[str, Any]) -> None:
    """Injects default LBO assumptions (e.g. 5.0x Entry/Exit multiples) if not extracted."""
    # 1. Entry Multiple
    entry_row = _find_row_by_label(ws, ["Purchase Assumptions", "Entry Multiple", "Entry Assumptions"])
    if entry_row:
        for r in range(entry_row, min(entry_row + 5, ws.max_row)):
            for c in range(1, 10):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and "multiple" in v.lower():
                    target_cell = ws.cell(row=r+1, column=c)
                    # Override dummy values or empty cells with 5.0x
                    if not getattr(target_cell, 'has_formula', False) and not str(target_cell.value).startswith('='):
                        target_cell.value = 5.0
                    break

    # 2. Exit Multiple
    exit_row = _find_row_by_label(ws, ["Exit Assumptions", "Exit Multiple"])
    if exit_row:
        for r in range(exit_row, min(exit_row + 5, ws.max_row)):
            for c in range(1, 10):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and "multiple" in v.lower():
                    target_cell = ws.cell(row=r+1, column=c)
                    if not getattr(target_cell, 'has_formula', False) and not str(target_cell.value).startswith('='):
                        target_cell.value = 5.0
                    break


def _inject_debt_profile(ws, data: Dict[str, Any]) -> None:
    """Injects extracted debt profile into Financing, Interest, and Amortization blocks."""
    dp = data.get("debt_profile", {}).get("facilities", [])
    if not isinstance(dp, list) or not dp:
        return
        
    for fac in dp:
        name = fac.get("name")
        if not name: continue
        bal = fac.get("balance")
        rate = fac.get("interest_rate_percent")
        amort = fac.get("amortization_per_year")
        
        name_lower = name.lower()
        # Clean "term loan a" -> "term loan" to match template
        if "term loan" in name_lower: name_lower = "term loan"
        elif "revolver" in name_lower: name_lower = "revolver"
        
        # 1. Update Sources Block (cols 2 and 3)
        sources_header = _find_row_by_label(ws, ["Financing Inputs", "Sources"])
        if sources_header:
            for r in range(sources_header, min(sources_header + 10, ws.max_row)):
                v = str(ws.cell(row=r, column=2).value or "").lower()
                if name_lower in v:
                    if bal is not None:
                        if not getattr(ws.cell(row=r, column=3), 'has_formula', False):
                            ws.cell(row=r, column=3).value = float(bal)
                    break

        # 2. Update Interest block (cols 11, 25, 26 and rates 27, 28)
        # Often Interest Inputs logic is around row 24-29
        int_header = _find_row_by_label(ws, ["Interest Inputs"])
        if int_header:
            for r in range(int_header, min(int_header + 10, ws.max_row)):
                v11 = str(ws.cell(row=r, column=11).value or "").lower()
                v25 = str(ws.cell(row=r, column=25).value or "").lower()
                v26 = str(ws.cell(row=r, column=26).value or "").lower()
                if name_lower in v11 or name_lower in v25 or name_lower in v26:
                    if rate is not None:
                        rate_dec = float(rate) / 100.0 if float(rate) > 1.0 else float(rate)
                        ws.cell(row=r, column=28).value = rate_dec
                        ws.cell(row=r, column=27).value = 0.0 # Clear spread
                    break
                    
        # 3. Update Amortization block
        amor_header = _find_row_by_label(ws, ["Amortization Inputs"])
        if amor_header:
            for r in range(amor_header, min(amor_header + 7, ws.max_row)):
                v11 = str(ws.cell(row=r, column=11).value or "").lower()
                v25 = str(ws.cell(row=r, column=25).value or "").lower()
                v26 = str(ws.cell(row=r, column=26).value or "").lower()
                if name_lower in v11 or name_lower in v25 or name_lower in v26:
                    if amort is not None:
                        ws.cell(row=r, column=27).value = float(amort)
                    break


def _clear_template_inputs(ws, year_blocks: List[Dict[str, Any]], row_map: Dict[str, Optional[int]]) -> None:
    cols: List[int] = []
    for b in year_blocks:
        bc = b.get("cols")
        if isinstance(bc, list):
            for c in bc:
                if isinstance(c, int):
                    cols.append(c)
    cols = sorted(set(cols))
    if not cols:
        return

    start_row, end_row = _infer_model_row_range(ws, row_map)
    if start_row is None or end_row is None:
        return

    for r in range(start_row, end_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            # Overwrite formulas too — we will write real numbers
            if isinstance(v, str) and v.startswith("="):
                cell.value = None
                continue
            if isinstance(v, (int, float)):
                cell.value = None
                continue
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    continue
                if s in {"-", "—", "–", "n/a", "na", "none"}:
                    cell.value = None
                    continue
                if any(ch.isdigit() for ch in s) and not re.search(r"[A-Za-z]", s):
                    cell.value = None


def _erase_excel_errors(ws) -> None:
    """Sweep entire sheet and replace any formula-error cells with '-'."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str):
                s = v.strip()
                # Replace Excel error strings
                if s in {"#DIV/0!", "#N/A", "#REF!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}:
                    cell.value = "-"
                    continue
                # Replace hanging formulas that survived
                if s.startswith("="):
                    cell.value = "-"
                    continue
                # Replace #### (column-too-narrow marker stored as string)
                if s.startswith("####"):
                    cell.value = "-"


def _pick_target_sheets(wb) -> list:
    """Return all worksheets; the caller will skip those without year blocks."""
    return [ws for ws in wb.worksheets]


def _sheet_contains_text(ws, needle: str) -> bool:
    needle = needle.lower()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400), min_col=1, max_col=min(ws.max_column, 120)):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and needle in v.lower():
                return True
    return False


def _detect_template_unit_scale(ws) -> float:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), min_col=1, max_col=min(ws.max_column, 30)):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            s = v.lower().replace(" ", "")
            if "inthousands" in s or "usdinthousands" in s or "$inthousands" in s or "$'000" in s or "$000s" in s:
                return 1000.0
            if "inmillions" in s or "usdinmillions" in s or "$inmillions" in s or "$'000,000" in s:
                return 1_000_000.0

    dvs = getattr(ws, "data_validations", None)
    dv_list = getattr(dvs, "dataValidation", None) if dvs is not None else None
    if isinstance(dv_list, list):
        for dv in dv_list:
            prompt = getattr(dv, "prompt", None)
            if not isinstance(prompt, str):
                continue
            s = prompt.lower().replace(" ", "")
            if "inthousands" in s or "usdinthousands" in s or "$inthousands" in s or "$'000" in s or "$000s" in s:
                return 1000.0
            if "inmillions" in s or "usdinmillions" in s or "$inmillions" in s or "$'000,000" in s:
                return 1_000_000.0
    return 1_000_000.0


def _parse_year_header(value: Any) -> Tuple[Optional[int], Optional[str]]:
    import datetime
    if isinstance(value, datetime.datetime):
        return value.year, "a"  # Treat historical dates natively as 'actual'
        
    if not isinstance(value, str):
        return None, None
    s = value.strip()
    
    # Try 20xx Suffix
    m = re.search(r"\b(20\d{2})\s*([aepfrm])?\b", s, re.IGNORECASE)
    if m:
        return int(m.group(1)), m.group(2)
        
    # Try FYxx Suffix
    m = re.search(r"FY\s*(\d{2})\s*([aepfrm])?\b", s, re.IGNORECASE)
    if m:
        return 2000 + int(m.group(1)), m.group(2)
        
    return None, None


def _detect_year_columns(ws) -> Dict[int, int]:
    best = (0, {})
    for r in range(1, min(ws.max_row, 120) + 1):
        found: Dict[int, int] = {}
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=r, column=c).value
            year, _ = _parse_year_header(v)
            if year is None:
                continue
            found.setdefault(year, c)
        if len(found) > best[0]:
            best = (len(found), found)
    return best[1]


def _detect_year_blocks(ws) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []

    for label, role in (
        ("Actual", "actual"),
        ("Atar Projections", "projection"),
        ("Management Projections", "management"),
        ("Projections", "projection"),
    ):
        blocks.extend(_detect_year_blocks_by_anchor(ws, label, role))

    # Generic detection to catch blocks missed by anchors
    generic_blocks = _detect_year_blocks_generic(ws)
    
    # Track occupied cells to avoid duplicates
    occupied_cells = set()
    for b in blocks:
        r = b.get("header_row")
        for c, _, _ in b.get("cells", []):
            occupied_cells.add((r, c))

    for gb in generic_blocks:
        r = gb.get("header_row")
        # Check if this block overlaps significantly with existing blocks
        # If any cell overlaps, we skip it (assuming anchors are more accurate)
        overlap = False
        for c, _, _ in gb.get("cells", []):
            if (r, c) in occupied_cells:
                overlap = True
                break
        
        if not overlap:
            blocks.append(gb)

    if blocks:
        for b in blocks:
            b["year_cols"] = {
                int(y): int(c)
                for c, y, _ in b.get("cells", [])
                if isinstance(c, int) and isinstance(y, int)
            }
        return blocks

    return []


def _detect_all_atar_blocks(ws) -> List[Dict[str, Any]]:
    atar_blocks = []
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, min(ws.max_column, 200) + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == "atar projections":
                start_col = c
                end_col = c
                for merged in ws.merged_cells.ranges:
                    if merged.min_row <= r <= merged.max_row and merged.min_col <= c <= merged.max_col:
                        start_col = merged.min_col
                        end_col = merged.max_col
                        break
                atar_blocks.append({
                    "header_row": r + 1,
                    "start_col": start_col,
                    "end_col": end_col,
                    "role": "projection"
                })
    return atar_blocks


def _detect_year_blocks_by_anchor(ws, label: str, role: str) -> List[Dict[str, Any]]:
    out_blocks: List[Dict[str, Any]] = []
    label_l = label.lower()
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, min(ws.max_column, 200) + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            if v.strip().lower() != label_l:
                continue
            header_row = r + 1
            # _scan_year_row_and_split returns a LIST of blocks now
            found_blocks = _scan_year_row_and_split(ws, header_row, c, role)
            if found_blocks:
                out_blocks.extend(found_blocks)
    return out_blocks


def _scan_year_row_and_split(ws, header_row: int, start_col: int, default_role: str) -> List[Dict[str, Any]]:
    cells: List[Tuple[int, int, str]] = []
    empty_streak = 0
    max_col = min(ws.max_column, start_col + 40)

    # 1. Collect all contiguous year-like cells
    for c in range(start_col, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        year, suffix = _parse_year_header(v)
        
        if year is None:
            empty_streak += 1
            if cells and empty_streak >= 3:
                break
            continue
            
        empty_streak = 0
        suffix = (suffix or "").strip().lower()
        cells.append((c, year, suffix))

    if not cells:
        return []

    # 2. Group by inferred role
    blocks: List[Dict[str, Any]] = []
    current_cells: List[Tuple[int, int, str]] = []
    current_role: Optional[str] = None
    prev_col: Optional[int] = None

    for c, year, suffix in cells:
        raw = ws.cell(row=header_row, column=c).value
        inferred = _infer_year_cell_role(raw, suffix)
        
        # If inferred is 'projection' but we are in a 'management' block, 
        # check if it's just 'r' or 'e' suffix vs 'm'.
        # For now, let's trust _infer_year_cell_role strictly.
        
        # Special case: if _infer_year_cell_role returns 'projection' (default),
        # but the default_role is 'actual' or 'management', AND suffix is missing, 
        # maybe we should inherit default_role?
        if inferred == "projection" and not suffix:
             # If no suffix, trust the anchor's default_role
             # But verify against "Actual" vs "Management" anchor logic
             if default_role in ("actual", "management"):
                 inferred = default_role

        if not current_cells:
            current_cells = [(c, year, suffix)]
            current_role = inferred
            prev_col = c
            continue

        if inferred == current_role and prev_col is not None and c == prev_col + 1:
            current_cells.append((c, year, suffix))
            prev_col = c
        else:
            # Finish current block
            if current_cells:
                blocks.append(_build_year_block(ws, header_row, current_cells))
            
            # Start new
            current_cells = [(c, year, suffix)]
            current_role = inferred
            prev_col = c

    if current_cells:
        blocks.append(_build_year_block(ws, header_row, current_cells))

    return blocks


def _detect_year_blocks_generic(ws) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []

    for r in range(1, min(ws.max_row, 200) + 1):
        matches: List[Tuple[int, int, str]] = []
        for c in range(1, min(ws.max_column, 200) + 1):
            v = ws.cell(row=r, column=c).value
            year, suffix = _parse_year_header(v)
            if year is None:
                continue
            suffix = (suffix or "").strip().lower()
            matches.append((c, year, suffix))
        if len(matches) < 3:
            continue

        matches.sort(key=lambda x: x[0])

        current_cells: List[Tuple[int, int, str]] = []
        current_role: Optional[str] = None
        prev_col: Optional[int] = None

        for c, year, suffix in matches:
            raw = ws.cell(row=r, column=c).value
            role = _infer_year_cell_role(raw, suffix)

            if not current_cells:
                current_cells = [(c, year, suffix)]
                current_role = role
                prev_col = c
                continue

            if role == current_role and prev_col is not None and c == prev_col + 1:
                current_cells.append((c, year, suffix))
                prev_col = c
            else:
                if current_cells:
                    cols = [cc for cc, _, _ in current_cells]
                    block = {
                        "header_row": r,
                        "cols": cols,
                        "cells": current_cells,
                        "role": current_role or "projection",
                    }
                    block["year_cols"] = {
                        int(y): int(cc)
                        for cc, y, _ in current_cells
                        if isinstance(cc, int) and isinstance(y, int)
                    }
                    blocks.append(block)

                current_cells = [(c, year, suffix)]
                current_role = role
                prev_col = c

        if current_cells:
            cols = [cc for cc, _, _ in current_cells]
            block = {
                "header_row": r,
                "cols": cols,
                "cells": current_cells,
                "role": current_role or "projection",
            }
            block["year_cols"] = {
                int(y): int(cc)
                for cc, y, _ in current_cells
                if isinstance(cc, int) and isinstance(y, int)
            }
            blocks.append(block)

    return blocks


def _infer_year_cell_role(raw: Any, suffix: str) -> str:
    raw_l = str(raw).strip().lower() if isinstance(raw, str) else ""
    suf = (suffix or "").strip().lower()

    if "actual" in raw_l:
        return "actual"
    if "management" in raw_l or "mgmt" in raw_l:
        return "management"
    if "projection" in raw_l or "projections" in raw_l or "proj" in raw_l or "atar" in raw_l:
        return "projection"

    if suf == "a":
        return "actual"
    if suf in {"m"}:
        return "management"
    if suf in {"p", "e"}:
        return "projection"

    return "projection"


def _build_year_block(ws, header_row: int, cells: List[Tuple[int, int, str]]) -> Dict[str, Any]:
    role = "projection"
    role_counts = {"actual": 0, "management": 0, "projection": 0}
    cols = []

    for c, _, suffix in cells:
        cols.append(c)
        raw = ws.cell(row=header_row, column=c).value
        raw_l = str(raw).lower() if raw is not None else ""
        
        # Simple heuristic based on suffix and content
        if "m" == suffix:
            role_counts["management"] += 1
        elif "a" == suffix:
            role_counts["actual"] += 1
        elif suffix in ("p", "e"):
            role_counts["projection"] += 1
        elif "management" in raw_l:
             role_counts["management"] += 1
        elif "actual" in raw_l:
             role_counts["actual"] += 1
        else:
            role_counts["projection"] += 1

    role = max(role_counts.items(), key=lambda kv: kv[1])[0]

    return {
        "header_row": header_row,
        "cols": cols,
        "cells": cells,
        "role": role
    }


def _infer_model_row_range(ws, row_map: Dict[str, Optional[int]]) -> Tuple[Optional[int], Optional[int]]:
    rows = [r for r in row_map.values() if isinstance(r, int)]
    if not rows:
        return None, None

    start = row_map.get("net_revenue") or min(rows)
    
    # We want the bottom-most row that matches any of our end markers
    end_markers = ["Revolver Balance", "Cash Available for Revolver", "Total Debt Service", "Debt Service"]
    
    found_ends = []
    for marker in end_markers:
        r = _find_row_by_label(ws, marker)
        if r:
            found_ends.append(r)
            
    end = max(found_ends) if found_ends else max(rows)
    
    if start and end and end >= start:
        return int(start), int(end)
    return int(min(rows)), int(max(rows))


def _infer_base_year(data: Dict[str, Any]) -> int:
    rev = data.get("revenue", {})
    years: List[int] = []
    if isinstance(rev, dict):
        hist = rev.get("history")
        if isinstance(hist, list):
            for it in hist:
                if not isinstance(it, dict):
                    continue
                y = _parse_year(it.get("period"))
                if y is not None:
                    years.append(int(y))
        present = rev.get("present")
        if isinstance(present, dict):
            y = _parse_year(present.get("period"))
            if y is not None:
                years.append(int(y))

    if years:
        return max(years)

    now_year = time.localtime().tm_year
    fallback_years = _collect_years_from_data(data)
    if fallback_years:
        past = [y for y in fallback_years if y <= now_year]
        return max(past) if past else max(fallback_years)
    return now_year


def _collect_years_from_data(data: Dict[str, Any]) -> List[int]:
    years: List[int] = []

    rev = data.get("revenue", {})
    if isinstance(rev, dict):
        for it in _collect_revenue_items(data):
            if not isinstance(it, dict):
                continue
            y = _parse_year(it.get("period"))
            if y is not None:
                years.append(y)

    profit = data.get("profit_metrics", {})
    if isinstance(profit, dict):
        for items in profit.values():
            if not isinstance(items, list):
                continue
            for it in items:
                if not isinstance(it, dict):
                    continue
                y = _parse_year(it.get("period"))
                if y is not None:
                    years.append(y)

    tale = data.get("tale_of_the_tape", {})
    if isinstance(tale, dict):
        for metric in tale.values():
            if not isinstance(metric, dict):
                continue
            yw = metric.get("year_wise")
            if not isinstance(yw, dict):
                continue
            for k in yw.keys():
                y = _parse_year(k)
                if y is not None:
                    years.append(y)

    fcf = data.get("free_cash_flow", {})
    if isinstance(fcf, dict):
        hist = fcf.get("historical")
        if isinstance(hist, dict):
            for k in hist.keys():
                y = _parse_year(k)
                if y is not None:
                    years.append(y)
        forecast = fcf.get("forecast_next_5_years")
        if isinstance(forecast, dict):
            for k in forecast.keys():
                y = _parse_year(k)
                if y is not None:
                    years.append(y)

    years = sorted(set(int(y) for y in years if isinstance(y, int)))
    return years


def _collect_revenue_items(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    rev = data.get("revenue", {})
    if not isinstance(rev, dict):
        return []
    out: List[Dict[str, Any]] = []
    for it in rev.get("history", []) if isinstance(rev.get("history"), list) else []:
        if isinstance(it, dict):
            out.append(it)
    present = rev.get("present")
    if isinstance(present, dict) and present:
        out.append(present)
    for it in rev.get("future", []) if isinstance(rev.get("future"), list) else []:
        if isinstance(it, dict):
            out.append(it)
    return out


def _get_actual_years_from_data(data: Dict[str, Any], base_year: int) -> List[int]:
    years = set()
    items = []
    
    # Collect from multiple sources just in case
    rev = data.get("revenue", {})
    if isinstance(rev, dict):
        items.extend(rev.get("history", []) if isinstance(rev.get("history"), list) else [])
        if isinstance(rev.get("present"), dict) and rev.get("present"):
            items.append(rev.get("present"))
            
    profit = data.get("profit_metrics", {})
    if isinstance(profit, dict):
        for metric_list in profit.values():
            if isinstance(metric_list, list):
                items.extend(metric_list)
                
    for item in items:
        if not isinstance(item, dict):
            continue
        period = str(item.get("period", "")).strip().upper()
        y = _parse_year(period)
        if y is None:
            continue
            
        suffix_indicators = ["B", "F", "E", "R", "M"]
        if period.endswith("A"):
            years.add(y)
        elif not any(period.endswith(s) for s in suffix_indicators):
            if y <= base_year:
                years.add(y)
                
    return sorted(list(years))


def _rewrite_year_header_row(ws, block: Dict[str, Any], base_year: int, actual_years: list = None) -> None:
    header_row = block.get("header_row")
    cols = block.get("cols")
    role = block.get("role")
    if not isinstance(header_row, int) or not isinstance(cols, list) or not cols:
        return
    if not isinstance(base_year, int):
        return

    n = len([c for c in cols if isinstance(c, int)])
    if n <= 0:
        return

    years = []
    if role == "actual":
        if actual_years:
            if len(actual_years) >= n:
                years = actual_years[-n:]
            else:
                pad = n - len(actual_years)
                first_actual = actual_years[0]
                years = list(range(first_actual - pad, first_actual)) + actual_years
        else:
            years = list(range(base_year - (n - 1), base_year + 1))
    else:
        years = list(range(base_year + 1, base_year + n + 1))

    for idx, c in enumerate(cols):
        if not isinstance(c, int) or idx >= len(years):
            continue
        cell = ws.cell(row=header_row, column=c)
        raw = cell.value
        _, suffix = _parse_year_header(raw)
        
        if not suffix:
            if role == "actual":
                suffix = "A"
            elif role == "management":
                suffix = "M"
            else:
                suffix = "R"
        else:
            suffix = suffix.upper()

        sep = " " if isinstance(raw, str) and " " in raw.strip() else ""
        cell.value = f"{years[idx]}{sep}{suffix}"


def _find_row_by_label(ws, labels: Any, min_row: int = 1, max_row: Optional[int] = None) -> Optional[int]:
    if isinstance(labels, str):
        labels = [labels]
    
    normalized_labels = {str(l).strip().lower() for l in labels if l}
    end = max_row if max_row is not None else min(ws.max_row, 400)
    
    for r in range(min_row, end + 1):
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                s = v.strip().lower()
                if s in normalized_labels:
                    return r
    return None


def _detect_row_map(ws) -> Dict[str, Optional[int]]:
    # Sequential bounding to prevent hooking duplicate labels out of order
    net_rev = _find_row_by_label(ws, ["Net Revenue", "Revenue", "Total Revenue", "Net Sales", "Sales"])
    cogs = _find_row_by_label(ws, ["COGS", "Cost of Goods Sold", "Cost of Sales", "Direct Costs"], min_row=net_rev or 1)
    gp = _find_row_by_label(ws, ["Gross Profit", "Gross Margin $", "Gross Income"], min_row=net_rev or 1)
    opex = _find_row_by_label(ws, ["Operating Expense", "Operating Expenses", "OpEx", "SG&A", "Total Operating Expenses"], min_row=gp or 1)
    ebitda = _find_row_by_label(ws, ["EBITDA", "Reported EBITDA", "Adjusted EBITDA (reported)", "Operating EBITDA"], min_row=opex or 1)
    
    adj_ebitda = _find_row_by_label(ws, ["PF Adj. EBITDA", "Adjusted EBITDA", "Adj. EBITDA"], min_row=ebitda or 1)
    fcf = _find_row_by_label(ws, ["Free Cash Flow", "FCF", "Unlevered Free Cash Flow"], min_row=adj_ebitda or 1)
    
    return {
        "net_revenue": net_rev,
        "cogs": cogs,
        "gross_profit": gp,
        "gross_margin": _find_row_by_label(ws, ["% Margin", "Gross Margin %", "Gross Profit Margin"], min_row=gp or 1),
        "operating_expense": opex,
        "ebitda": ebitda,
        "ebitda_margin": _find_second_percent_margin_row(ws, after_label=(ebitda or 1)) if ebitda else None,
        "pf_adjustments": _find_row_by_label(ws, ["PF Adjustments", "EBITDA Adjustments", "Adjustments"], min_row=ebitda or 1),
        "adj_ebitda": adj_ebitda,
        "adj_ebitda_margin": _find_second_percent_margin_row(ws, after_label=(adj_ebitda or 1)) if adj_ebitda else None,
        "capex": _find_row_by_label(ws, ["Capex", "Capital Expenditures", "Capital Expenditure", "Additions to PPE"], min_row=adj_ebitda or 1),
        "change_in_wc": _find_row_by_label(ws, ["Change in WC", "Change in Working Capital", "Working Capital Change", "(Increase)/Decrease in WC"], min_row=adj_ebitda or 1),
        "one_time_cost": _find_row_by_label(ws, ["1x Costs", "One-time Costs", "Non-recurring", "EBITDA Normalizations", "1x Costs"]),
        "free_cash_flow": fcf,
        # Debt service section
        "interest": _find_row_by_label(ws, ["Interest", "Interest Expense", "Total Interest", "Interest Subtotal"], min_row=fcf or 1),
        "amortization": _find_row_by_label(ws, ["Amortization", "Debt Amortization", "Principal Amortization", "Amortization Subtotal"], min_row=fcf or 1),
        "total_debt_service": _find_row_by_label(ws, ["Total Debt Service", "Debt Service", "Total Debt Service Subtotal"], min_row=fcf or 1),
        "fcf_after_debt": _find_row_by_label(ws, ["FCF After Debt Service", "FCF After Debt", "Cash After Debt Service"], min_row=fcf or 1),
        "cash_avail_revolver": _find_row_by_label(ws, ["Cash Available for Revolver", "Cash Available", "Cash Available Revolver"], min_row=fcf or 1),
        "revolver_draw": _find_row_by_label(ws, ["Revolver Draw", "Revolver Draw / Repayment", "Revolver Drawing"], min_row=fcf or 1),
        "revolver_balance": _find_row_by_label(ws, ["Revolver Balance", "Revolver Outstanding", "Revolver Ending Balance"], min_row=fcf or 1),
        "remaining_cash": _find_row_by_label(ws, ["Remaining Cash", "Net Cash", "Cash After Revolver"], min_row=fcf or 1),
    }


def _find_second_percent_margin_row(ws, after_label: Any) -> Optional[int]:
    if isinstance(after_label, int):
        start = after_label
    else:
        start = _find_row_by_label(ws, after_label)
        if not start:
            return None
    
    target_labels = {"% margin", "margin %", "gross margin %", "ebitda margin %", "margin"}
    
    for r in range(start + 1, min(start + 8, ws.max_row) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                s = v.strip().lower()
                if s in target_labels:
                    return r
    return None


def _parse_amount(value: Any, template_scale: float) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s or s == "-":
        return None

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()
    if s.startswith("-"):
        neg = True
        s = s[1:].strip()

    s_clean = s.replace("$", "").replace(",", "").strip()

    mult = 1.0
    s_l = s.lower()
    if "billion" in s_l or s_l.endswith("b"):
        mult = 1_000_000_000.0
        s_clean = re.sub(r"(?i)billion|\\bb\\b", "", s_clean).strip()
    elif "million" in s_l or s_l.endswith("m"):
        mult = 1_000_000.0
        s_clean = re.sub(r"(?i)million|\\bm\\b", "", s_clean).strip()
    elif "thousand" in s_l or s_l.endswith("k"):
        mult = 1_000.0
        s_clean = re.sub(r"(?i)thousand|\\bk\\b", "", s_clean).strip()

    try:
        parts = re.findall(r"-?\\d+(?:\\.\\d+)?", s_clean)
        if not parts:
            return None
        if len(parts) == 1:
            num = float(parts[0])
        else:
            nums = [float(p) for p in parts]
            num = sum(nums) / len(nums)
    except Exception:
        return None

    out = (num * mult) / template_scale
    return -out if neg else out


def _value_mentions_scale(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip().lower()
    if not s:
        return False
    s_clean = re.sub(r"[\\s,$]", "", s)
    return any(tok in s for tok in ("billion", "million", "thousand")) or s_clean.endswith(("b", "m", "k"))


def _unit_multiplier_from_unit(unit: Any) -> float:
    if unit is None:
        return 1.0
    s = str(unit).strip().lower()
    if not s:
        return 1.0
    s_clean = re.sub(r"[\\s,$]", "", s)
    if "billion" in s_clean or s_clean in {"b", "bn"} or s_clean.endswith("b"):
        return 1_000_000_000.0
    if "million" in s_clean or s_clean in {"m", "mm"} or s_clean.endswith("m"):
        return 1_000_000.0
    if "thousand" in s_clean or s_clean in {"k"} or s_clean.endswith("k"):
        return 1_000.0
    return 1.0


def _prepare_tale_projections(
    base_year: int,
    historical_data: Dict[int, float],
    revenue_atar: Dict[int, float],
    revenue_mgmt: Dict[int, float],
    metric_type: str
) -> tuple:
    """
    Prepare Tale of the Tape projections for both Atar and Management scenarios.
    
    Args:
        base_year: The last actual year
        historical_data: Historical values from extraction
        revenue_atar: Atar revenue projections
        revenue_mgmt: Management revenue projections
        metric_type: Type of metric ("capex", "wc", "one_time")
    
    Returns:
        Tuple of (atar_dict, mgmt_dict) with year-wise projections
    """
    atar_data = dict(historical_data)
    mgmt_data = dict(historical_data)
    
    # Determine projection years (next 5 years after base_year)
    projection_years = list(range(base_year + 1, base_year + 6))
    
    if metric_type == "capex":
        # CAPEX projection logic
        # Conservative approach: Use % of revenue or last known value
        
        # Find last known CAPEX value
        hist_years = sorted([y for y in historical_data.keys() if y <= base_year and historical_data.get(y)])
        if hist_years:
            last_capex = historical_data[hist_years[-1]]
            
            # Calculate CAPEX as % of revenue if possible
            revenue_at_year = revenue_atar.get(hist_years[-1]) or revenue_mgmt.get(hist_years[-1])
            if revenue_at_year and revenue_at_year != 0:
                capex_pct = abs(last_capex) / revenue_at_year
                capex_pct = min(max(capex_pct, 0.02), 0.10)  # Cap between 2% and 10%
            else:
                capex_pct = 0.04  # Default 4% of revenue
        else:
            capex_pct = 0.04  # Default 4% of revenue
            last_capex = None
        
        # Project CAPEX for future years
        for year in projection_years:
            if year not in atar_data:
                rev_atar = revenue_atar.get(year)
                if rev_atar:
                    atar_data[year] = -(rev_atar * capex_pct)  # Negative for cash outflow
                elif last_capex is not None:
                    atar_data[year] = last_capex
            
            if year not in mgmt_data:
                rev_mgmt = revenue_mgmt.get(year)
                if rev_mgmt:
                    mgmt_data[year] = -(rev_mgmt * capex_pct)  # Same % for management
                elif last_capex is not None:
                    mgmt_data[year] = last_capex
    
    elif metric_type == "wc":
        # Working Capital Change projection
        # Formula: -20% of revenue growth (already in template, but we can pre-calculate)
        # Negative change in WC = cash outflow (growing business needs more working capital)
        
        for year in projection_years:
            if year not in atar_data:
                # For Atar: calculate based on revenue growth
                prev_year_rev = revenue_atar.get(year - 1)
                curr_year_rev = revenue_atar.get(year)
                if prev_year_rev and curr_year_rev:
                    rev_growth = curr_year_rev - prev_year_rev
                    atar_data[year] = -(rev_growth * 0.20)  # 20% of revenue growth
            
            if year not in mgmt_data:
                # For Management: calculate based on management revenue growth
                prev_year_rev = revenue_mgmt.get(year - 1)
                curr_year_rev = revenue_mgmt.get(year)
                if prev_year_rev and curr_year_rev:
                    rev_growth = curr_year_rev - prev_year_rev
                    mgmt_data[year] = -(rev_growth * 0.20)
    
    elif metric_type == "one_time":
        # One-time costs typically don't recur in projections
        # Set to 0 for future years unless document specifies otherwise
        for year in projection_years:
            if year not in atar_data:
                atar_data[year] = 0
            if year not in mgmt_data:
                mgmt_data[year] = 0
    
    return atar_data, mgmt_data


def _fill_future_projections(
    base_year: int,
    revenue: Dict[int, float],
    gross_profit: Dict[int, float],
    operating_income: Dict[int, float],
    opex: Dict[int, float],
    ebitda: Dict[int, float],
    mode: str = "atar"
) -> None:
    """
    Extends projections for Revenue, Gross Profit, and EBITDA for 5 years into the future.
    Uses simple CAGR or margin-based logic if data is missing.
    
    Args:
        base_year: The last actual year.
        revenue: Dictionary of revenue by year.
        gross_profit: Dictionary of gross profit by year.
        ebitda: Dictionary of EBITDA by year.
        mode: "atar" (conservative/AI) or "management" (optimistic).
    """
    # Determine how far to project (e.g. 5 years)
    target_years = range(base_year + 1, base_year + 6)
    
    # --- 1. Extend Revenue ---
    known_rev_years = sorted([y for y in revenue.keys() if y <= base_year])
    
    cagr = 0.05  # Default fallback
    
    if known_rev_years:
        last_rev_y = known_rev_years[-1]
        last_rev_val = revenue[last_rev_y]
        
        # Calculate recent CAGR
        if len(known_rev_years) >= 2:
            first_rev_y = known_rev_years[0]
            first_rev_val = revenue[first_rev_y]
            
            # Use a shorter window for CAGR if possible (e.g. last 3 years) to capture recent trends
            if len(known_rev_years) > 3:
                first_rev_y = known_rev_years[-3]
                first_rev_val = revenue[first_rev_y]

            if first_rev_val > 0 and last_rev_val > 0:
                years_diff = last_rev_y - first_rev_y
                if years_diff > 0:
                    raw_cagr = (last_rev_val / first_rev_val) ** (1 / years_diff) - 1
                    
                    if mode == "management":
                        # Management case: Optimistic
                        # Cap at reasonable high (e.g. +30%), floor at 0% (unless significantly negative trend)
                        cagr = max(0.02, min(0.30, raw_cagr * 1.1)) # 10% boost to historical CAGR
                        if raw_cagr < 0: cagr = 0.02 # Turnaround assumption for management
                    else:
                        # ATAR case: Conservative
                        # Cap at reasonable low (e.g. +15%), floor at -10%
                        cagr = max(-0.10, min(0.15, raw_cagr * 0.9)) # 10% haircut to historical CAGR
    else:
        # No history
        last_rev_val = 0
        last_rev_y = base_year
        if mode == "management":
             cagr = 0.10
        else:
             cagr = 0.05

    # Fill missing revenue
    # Note: We respect existing future values if present (e.g. extracted from document)
    # UNLESS they are significantly inconsistent or we want to override?
    # For now, we only fill GAPS.
    
    current_val = last_rev_val
    # Initialize current_val correctly for the loop if we have gaps
    
    # Fill gap years up to target_years if any
    gap_years = range(last_rev_y + 1, target_years[0]) if target_years and last_rev_y else []
    for y in gap_years:
        if y not in revenue:
            new_val = current_val * (1 + cagr)
            revenue[y] = new_val
            current_val = new_val
    
    for y in target_years:
        if y in revenue:
            current_val = revenue[y]
        else:
            prev_val = revenue.get(y - 1)
            if prev_val is None:
                 prev_val = current_val 
            
            if prev_val is not None and prev_val != 0:
                new_val = prev_val * (1 + cagr)
                revenue[y] = new_val
                current_val = new_val

    # --- 2. Extend Gross Profit ---
    known_gp_years = sorted([y for y in gross_profit.keys() if y <= base_year])
    
    avg_gm = 0.40 # Default
    margins = []
    
    if known_gp_years:
        recent_years = known_gp_years[-3:]
        for y in recent_years:
            r = revenue.get(y)
            g = gross_profit.get(y)
            if r and g and r != 0:
                margins.append(g / r)
        
        if margins:
            if mode == "management":
                # Optimistic: Use max margin or average + boost
                avg_gm = max(margins)
            else:
                # Conservative: Use average or min
                avg_gm = sum(margins) / len(margins)

    for y in target_years:
        if y in gross_profit:
            continue
        rev = revenue.get(y)
        if rev is not None:
            gross_profit[y] = rev * avg_gm

    # --- 3. Extend OpEx ---
    known_opex_years = sorted([y for y in opex.keys() if y <= base_year])
    avg_opex_margin = 0.25 # Default 25% of revenue
    opex_margins = []
    
    if known_opex_years:
        recent_years = known_opex_years[-3:]
        for y in recent_years:
            r = revenue.get(y)
            o = opex.get(y)
            if r and o and r != 0:
                opex_margins.append(abs(o) / r)
                
        if opex_margins:
            if mode == "management":
                avg_opex_margin = sum(opex_margins) / len(opex_margins) * 0.9 # Optimization
            else:
                avg_opex_margin = sum(opex_margins) / len(opex_margins)
                
    for y in target_years:
        if y in opex:
            continue
        rev = revenue.get(y)
        if rev is not None:
            opex[y] = rev * avg_opex_margin

    # --- 4. Extend EBITDA ---
    # According to new rule: For each projection year, EBITDA = Revenue - COGS - Opex
    # And Revenue - COGS = Gross Profit. So EBITDA = Gross Profit - Opex.
    for y in target_years:
        if y in ebitda:
            continue
        g = gross_profit.get(y)
        o = opex.get(y)
        if g is not None and o is not None:
            # We must subtract absolute OpEx, since OpEx might be stored as positive naturally
            ebitda[y] = g - abs(o)


def _extract_series_by_year(items: Any, template_scale: float) -> Dict[int, float]:
    if not isinstance(items, list):
        return {}

    out: Dict[int, float] = {}
    for it in items:
        if not isinstance(it, dict):
            continue
        period = it.get("period")
        y = _parse_year(period)
        if y is None:
            # print(f"DEBUG: Failed to parse year from period: '{period}'")
            continue
        raw_value = it.get("value")
        unit_mult = _unit_multiplier_from_unit(it.get("unit"))

        if isinstance(raw_value, (int, float)):
            val = (float(raw_value) * unit_mult) / template_scale
        else:
            val = _parse_amount(raw_value, template_scale)
            if val is None:
                # print(f"DEBUG: Failed to parse amount: '{raw_value}' for year {y}")
                continue
            if unit_mult != 1.0 and not _value_mentions_scale(raw_value):
                val = val * unit_mult

        out[y] = float(val)
    # print(f"DEBUG: Extracted series: {out}")
    return out


def _extract_tale_year_wise(data: Dict[str, Any], key: str, template_scale: float) -> Dict[int, float]:
    tale = data.get("tale_of_the_tape")
    if not isinstance(tale, dict):
        return {}
    metric = tale.get(key)
    if not isinstance(metric, dict):
        return {}
    year_wise = metric.get("year_wise")
    if not isinstance(year_wise, dict):
        return {}

    unit_mult = _unit_multiplier_from_unit(metric.get("unit"))
    out: Dict[int, float] = {}
    for year_label, obj in year_wise.items():
        y = _parse_year(year_label)
        if y is None:
            continue
        v = obj.get("value") if isinstance(obj, dict) else obj
        if isinstance(v, (int, float)):
            val = (float(v) * unit_mult) / template_scale
        else:
            val = _parse_amount(v, template_scale)
            if val is None:
                continue
            if unit_mult != 1.0 and not _value_mentions_scale(v):
                val = val * unit_mult
        out[y] = float(val)
    return out


def _extract_fcf(data: Dict[str, Any], template_scale: float) -> Tuple[Dict[int, float], Dict[int, float]]:
    fcf = data.get("free_cash_flow")
    if not isinstance(fcf, dict):
        return {}, {}

    hist = fcf.get("historical")
    out_hist: Dict[int, float] = {}
    default_unit_mult = 1.0
    rev = data.get("revenue", {})
    if isinstance(rev, dict):
        for it in rev.get("history", []) if isinstance(rev.get("history"), list) else []:
            if not isinstance(it, dict):
                continue
            u = it.get("unit")
            m = _unit_multiplier_from_unit(u)
            if m != 1.0:
                default_unit_mult = m
                break

    if isinstance(hist, dict):
        for year_label, obj in hist.items():
            y = _parse_year(year_label)
            if y is None:
                continue
            v = obj.get("value") if isinstance(obj, dict) else obj
            if isinstance(v, (int, float)):
                val = (float(v) * default_unit_mult) / template_scale
            else:
                val = _parse_amount(v, template_scale)
                if val is None:
                    continue
                if default_unit_mult != 1.0 and not _value_mentions_scale(v):
                    val = val * default_unit_mult
            out_hist[y] = float(val)

    forecast = fcf.get("forecast_next_5_years")
    out_forecast: Dict[int, float] = {}
    if isinstance(forecast, dict):
        for k, v in forecast.items():
            if k in {"base_year", "growth_rate_used", "methodology"}:
                continue
            y = _parse_year(k)
            if y is None:
                continue
            if isinstance(v, (int, float)):
                val = (float(v) * default_unit_mult) / template_scale
            else:
                val = _parse_amount(v, template_scale)
                if val is None:
                    continue
                if default_unit_mult != 1.0 and not _value_mentions_scale(v):
                    val = val * default_unit_mult
            out_forecast[y] = float(val)

    return out_hist, out_forecast


def _parse_year(period: Any) -> Optional[int]:
    if period is None:
        return None
    s = str(period).strip().upper()
    
    # Try 4-digit year first (with optional FY/CY prefix and optional suffixes)
    m = re.search(r"(?:FY|CY)?\s*(20\d{2})", s)
    if m:
        return int(m.group(1))

    # Try 2-digit year with prefix (e.g. FY26)
    m = re.search(r"(?:FY|CY)\s*(\d{2})", s)
    if m:
        return 2000 + int(m.group(1))
        
    # Try 2-digit year with suffix (e.g. 23E, 24A)
    m = re.search(r"\b(\d{2})\s*[AEPFMB]\b", s)
    if m:
        return 2000 + int(m.group(1))
        
    # print(f"DEBUG: _parse_year failed for '{s}'")
    return None


def _derive_cogs(revenue: Dict[int, float], gross_profit: Dict[int, float]) -> Dict[int, float]:
    out: Dict[int, float] = {}
    for y in set(revenue.keys()) & set(gross_profit.keys()):
        out[y] = revenue[y] - gross_profit[y]
    return out


def _derive_opex(gross_profit: Dict[int, float], operating_income: Dict[int, float]) -> Dict[int, float]:
    out: Dict[int, float] = {}
    for y in set(gross_profit.keys()) & set(operating_income.keys()):
        out[y] = gross_profit[y] - operating_income[y]
    return out


def _derive_adj_ebitda(ebitda: Dict[int, float], adjustments: Dict[int, float]) -> Dict[int, float]:
    out: Dict[int, float] = {}
    for y in set(ebitda.keys()) | set(adjustments.keys()):
        e = ebitda.get(y)
        a = adjustments.get(y, 0.0)
        if e is None:
            continue
        out[y] = e + a
    return out


def _write_line(
    ws, 
    row: Optional[int], 
    year_cols: Dict[int, int], 
    values: Dict[int, float], 
    force_negative: bool = False, 
    is_ebitda: bool = False,
    template_scale: Optional[float] = None
) -> None:
    if not row or not values:
        return
    for year, col in year_cols.items():
        if year not in values:
            continue
        cell = ws.cell(row=row, column=col)
            
        v = float(values[year])
        
        # Format the specific value but DO NOT wrap it in brackets yourself
        # just force mathematically negative, and let the excel format string handle UI.
        if force_negative and v > 0:
            v = -v
            
        cell.value = v
        
        if template_scale is not None:
            if template_scale >= 1_000_000.0:
                fmt = '"$"#,##0.0"M";[Red]("$"#,##0.0"M")'
            elif template_scale >= 1_000.0:
                fmt = '"$"#,##0.0,"M";[Red]("$"#,##0.0,"M")'
            else:
                fmt = '"$"#,##0.0,,"M";[Red]("$"#,##0.0,,"M")'
            cell.number_format = fmt
        elif is_ebitda:
            cell.number_format = '"$"#,##0.0;[Red]("$"#,##0.0)'


def _write_percent_line(
    ws,
    row: Optional[int],
    year_cols: Dict[int, int],
    numerator: Dict[int, float],
    denominator: Dict[int, float]
) -> None:
    if not row:
        return
    for year, col in year_cols.items():
        num = numerator.get(year)
        den = denominator.get(year)
        if num is None or den in (None, 0):
            continue
        cell = ws.cell(row=row, column=col)
        ratio = float(num) / float(den)
        if isinstance(cell.number_format, str) and "%" in cell.number_format:
            cell.value = ratio
        else:
            cell.value = ratio * 100.0


def _fuzzy_find_row_by_label(ws, target_name: str, min_row: int = 1, max_row: Optional[int] = None) -> Optional[int]:
    import difflib
    target = target_name.lower().strip()
    if not target: return None
    end = max_row if max_row is not None else min(ws.max_row, 400)
    best_match = None
    best_score = 0.8  # Threshold for fuzzy match

    for r in range(min_row, end + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                s = v.strip().lower()
                if not s: continue
                # Exact or substring match highly prioritized
                if target == s:
                    return r
                if target in s:
                    # e.g., "accounts receivable" in "accounts receivable, net"
                    return r
                # Fuzzy fallback
                ratio = difflib.SequenceMatcher(None, target, s).ratio()
                if ratio > best_score:
                    best_score = ratio
                    best_match = r
                    
    return best_match


def _write_balance_sheet_to_excel(ws, year_blocks: List[Dict[str, Any]], bs_data: Dict[str, Any], template_scale: float) -> None:
    if not isinstance(bs_data, dict):
        return
        
    actual_year_cols = {}
    for block in year_blocks:
        role = block.get("role", "projection")
        if role == "actual" or role == "projection":
            yc = block.get("year_cols", {})
            actual_year_cols.update(yc)
            
    if not actual_year_cols:
        return
        
    for section in ["assets", "liabilities", "equity"]:
        items = bs_data.get(section, [])
        if not isinstance(items, list):
            continue
            
        for item in items:
            name = item.get("item_name")
            values_list = item.get("values", [])
            if not name or not isinstance(values_list, list):
                continue
                
            row = _fuzzy_find_row_by_label(ws, name)
            if not row:
                continue
                
            # Treat array of period/values like extracted array
            series = _extract_series_by_year(values_list, template_scale)
            _write_line(ws, row, actual_year_cols, series, force_negative=False, template_scale=template_scale)
